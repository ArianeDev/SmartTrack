from .models import User, Sensor, Ambient, History
from .permission import IsAdm
from .serializer import SensorsSerializer, AmbientSerializer, HistorySerializer, UserSerializer, LoginSerializer
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import render
from io import BytesIO
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
import openpyxl
import pandas as pd

# Upload CSV file
# Sensors, Ambient and History
        
class UploadFile_Sensors(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAdm]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        if not file:
            return Response({"error": "Nenhum arquivo enviado"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)
            # Columns CSV
            columns = {"sensor", "mac_address", "unidade_medida", "longitude", "latitude", "status"}

            if not columns.issubset(df.columns):
                return Response({"error": "Colunas inválidas no arquivo CSV"}, status=status.HTTP_400_BAD_REQUEST)
            
            data = [
                # Insert in the database
                Sensor(
                    type_sensors=row['sensor'],
                    mac_address=row['mac_address'],
                    unit_measure=row['unidade_medida'],
                    longitude=row['longitude'],
                    latitude=row['latitude'],
                    status=row['status']
                )

                for _, row in df.iterrows()
            ]
            Sensor.objects.bulk_create(data)
            return Response({"message": "Arquivo CSV processado com sucesso"}, status=status.HTTP_201_CREATED)
    
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class UploadFile_Ambient(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAdm]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        if not file:
            return Response({"error": "Nenhum arquivo enviado"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)
            # Columns CSV
            columns = {"sig", "descricao", "ni", "responsavel"}

            if not columns.issubset(df.columns):
                return Response({"error": "Colunas inválidas no arquivo CSV"}, status=status.HTTP_400_BAD_REQUEST)
            
            data = [
                # Insert in the database
                Ambient(
                    sig=row['sig'],
                    description=row['descricao'],
                    ni=row['ni'],
                    responsible=row['responsavel']
                )

                for _, row in df.iterrows()
            ]
            Ambient.objects.bulk_create(data)
            return Response({"message": "Arquivo CSV processado com sucesso"}, status=status.HTTP_201_CREATED)
    
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class UploadFile_History(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAdm]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        if not file:
            return Response({"error": "Nenhum arquivo enviado"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)
            # Columns CSV
            columns = {"sensor", "ambiente", "valor", "timestamp"}

            if not columns.issubset(df.columns):
                return Response({"error": "Colunas inválidas no arquivo CSV"}, status=status.HTTP_400_BAD_REQUEST)
            
            data = []

            # Insert in the database
            for _, row in df.iterrows():
                try:
                    # verification if the sensor and ambient exist in database
                    sensor = Sensor.objects.get(id=row['sensor'])  
                    ambient = Ambient.objects.get(id=row['ambiente'])  

                    data.append(History(
                        sensor=sensor,  
                        ambient=ambient,  
                        value=row['valor'],
                        timestamp=row['timestamp']
                    ))
                except (Sensor.DoesNotExist, Ambient.DoesNotExist) as e:
                    print(f"Erro: {e}")

            History.objects.bulk_create(data)
            return Response({"message": "Arquivo CSV processado com sucesso"}, status=status.HTTP_201_CREATED)
    
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

# Export CSV file
# Sensors, Ambient and History

class ExportFile_Sensors(APIView):
    permission_classes = [IsAdm]

    def get_queryset(self):
        return Sensor.objects.all()

    def get(self, request):
        try:
            sensor_type = request.query_params.get('type_sensors')

            # Check if the type_sensors exists in the request
            if not sensor_type:
                return Response({"error": "Parâmetro 'type_sensors' é obrigatório"}, status=status.HTTP_400_BAD_REQUEST)
            
            sensors = self.get_queryset().filter(type_sensors=sensor_type)

            # Check if there are sensors with the specified type
            if not sensors.exists():
                return Response({"error": "Nenhum sensor encontrado com o tipo especificado"}, status=status.HTTP_404_NOT_FOUND)
            
            wb = openpyxl.Workbook() # create file
            ws = wb.active
            ws.title = sensor_type.capitalize() # title the file

            ws.append(["sensor", "mac_address", "unidade_medida", "latitude", "longitude", "status"])

            for sensor in sensors:
                ws.append([
                    sensor.type_sensors,
                    sensor.mac_address,
                    sensor.unit_measure,
                    sensor.longitude,
                    sensor.latitude,
                    sensor.status
                ])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"{sensor_type}.xlsx"

            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response
        
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ExportFile_Ambient(APIView):
    permission_classes = [IsAdm]

    def get_queryset(self):
        return Ambient.objects.all()

    def get(self, request):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ambients"

            ws.append(["SIG", "Descrição", "NI", "Responsável"])

            for ambient in self.get_queryset():
                ws.append([
                    ambient.sig,
                    ambient.description,
                    ambient.ni,
                    ambient.responsible
                ])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            response['Content-Disposition'] = 'attachment; filename="ambients.xlsx"'

            return response
        
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class ExportFile_History(APIView):
    permission_classes = [IsAdm]

    def get_queryset(self):
        return History.objects.all()

    def get(self, request):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "History"

            ws.append(["sensor", "ambiente", "valor", "timestamp"])

            for history in self.get_queryset():    
                ws.append([
                    history.sensor.mac_address,
                    history.ambient.description,  
                    history.value,
                    history.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            response['Content-Disposition'] = 'attachment; filename="history.xlsx"'

            return response
        
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

# Login

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def GetUser(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)

class Login(TokenObtainPairView):
    serializer_class = LoginSerializer

# User methods

class User_GET_POST(ListCreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_permissions(self):  
        print("get_permissions executado")
        if self.request.method == 'POST':
            return [AllowAny()] # allow any user to create a new user
        return [IsAdm()]

# Sensors methods

class SensorsPaginated(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 50

class Sensors_GET_POST(ListCreateAPIView):
    queryset = Sensor.objects.all()
    serializer_class = SensorsSerializer
    pagination_class = SensorsPaginated
    permission_classes = [IsAdm]

    def get_queryset(self):
        queryset = super().get_queryset()
        type_sensors = self.request.query_params.get('type_sensors')

        if type_sensors:
            queryset = queryset.filter(type_sensors=type_sensors)

        return queryset

class Sensors_GET_PUT_PATCH_DELETE(RetrieveUpdateDestroyAPIView):
    queryset = Sensor.objects.all()
    serializer_class = SensorsSerializer
    permission_classes = [IsAdm]

# Ambient methods

class Ambient_GET_POST(ListCreateAPIView):
    queryset = Ambient.objects.all()
    serializer_class = AmbientSerializer
    permission_classes = [IsAdm]

class Ambient_GET_PUT_PATCH_DELETE(RetrieveUpdateDestroyAPIView):
    queryset = Ambient.objects.all()
    serializer_class = AmbientSerializer
    permission_classes = [IsAdm]

# History methods
class HistoryPaginated(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 50 

class History_GET_POST(ListCreateAPIView):
    queryset = History.objects.all().order_by('-timestamp')
    serializer_class = HistorySerializer
    pagination_class = HistoryPaginated
    permission_classes = [IsAdm]

    def get_queryset(self):
        queryset = super().get_queryset()
        id_history = self.request.query_params.get('id')
        id_sensor = self.request.query_params.get('sensor')
        id_ambient = self.request.query_params.get('ambient')
        type_sensors = self.request.query_params.get('type_sensors')
        date = self.request.query_params.get('timestamp')
        VAILID_SENSOR = {"temperatura", "umidade", "luminosidade", "contador"}

        # filter by id history
        try:
            if id_history:
                queryset = queryset.filter(id=id_history)
        except Exception as e:
            raise ValidationError({"error": "ID de histórico inválido."})

        # filter by sensor id
        try:
            if id_sensor:
                queryset = queryset.filter(sensor=id_sensor)
        except Exception as e:
            raise ValidationError({"error": "ID de ambiente inválido."})
        
        # filter by ambient id
        try:
            if id_ambient:
                queryset = queryset.filter(ambient=id_ambient)
        except Exception as e:
            raise ValidationError({"error": "ID de ambiente inválido."})

        # filter by type of sensor
        try:
            if type_sensors:
                if type_sensors.lower() not in VAILID_SENSOR:
                    raise ValidationError({"error": "Tipo de sensor inválido. Use: temperatura, umidade, luminosidade ou contador."})
                
                queryset = queryset.filter(sensor__type_sensors=type_sensors)
        except ValidationError as e:
            raise e
        
        # filter by date
        try:
            if date:
                queryset = queryset.filter(timestamp__date=date)
        except Exception as e:
            raise ValidationError({"error": "Data inválida. Use o formato YYYY-MM-DD."})

        if queryset:
            return queryset
        else:
            raise ValidationError({"message": "Não foi encontrado nenhum registro.", "status": status.HTTP_404_NOT_FOUND})

class History_GET_PUT_PATCH_DELETE(RetrieveUpdateDestroyAPIView):
    queryset = History.objects.all()
    serializer_class = HistorySerializer
    permission_classes = [IsAdm]